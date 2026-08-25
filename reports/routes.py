"""Reports routes.

Per-department report filtered by month / date, returning client, show, shot,
mandays and client feedback for each shot.
"""

import os
from datetime import date, datetime

from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from auth.middleware import token_required
from common.db_utils import run_query, to_iso
from common.http import failure, success

reports_bp = Blueprint("reports", __name__)

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports", "reports")

_HEADER_FILL = PatternFill("solid", fgColor="3EBA02")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_CELL_ALIGN = Alignment(vertical="center")


def _safe(value):
    return "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in str(value)).strip("_") or "report"


def _parse_date_range(start_date_raw, end_date_raw):
    if not start_date_raw or not end_date_raw:
        return None, None
    try:
        start_date = datetime.strptime(start_date_raw, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_raw, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None, None
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return start_date, end_date


# Progress bucket used by the mandays bar chart on the report screen.
_COMPLETED_STATUSES = {"Completed", "Approved", "Approved Internal"}
_IN_PROGRESS_STATUSES = {"WIP", "Bidding", "Bids Received", "Delivered"}
_IN_PROGRESS_ARTIST = {
    "In Progress",
    "Awaiting QC",
    "WIP Completed",
    "Render & Upload Completed",
}


def _progress_of(row):
    """Bucket a shot into Completed / In Progress / Remaining for the chart."""
    status = (row.get("status") or "").strip()
    supervisor = (row.get("supervisor_status") or "").strip()
    artist = (row.get("artist_status") or "").strip()
    if status in _COMPLETED_STATUSES or supervisor == "Approved":
        return "Completed"
    if status in _IN_PROGRESS_STATUSES or artist in _IN_PROGRESS_ARTIST:
        return "In Progress"
    return "Remaining"


def _report_rows(
    department,
    specific_date=None,
    start_date=None,
    end_date=None,
    month=None,
    year=None,
):
    today = date.today()
    # Department may be a comma-separated list (multi-department shots).
    dept_parts = [d.strip() for d in (department or "").split(",") if d.strip()]
    if len(dept_parts) == 1:
        clauses = ["s.department = %s"]
        params = [dept_parts[0]]
    else:
        dept_clause = " OR ".join(["FIND_IN_SET(%s, s.department)"] * len(dept_parts))
        clauses = [f"({dept_clause})"]
        params = dept_parts[:]

    if start_date and end_date:
        clauses.append("DATE(s.allocated_date) BETWEEN %s AND %s")
        params.extend([start_date, end_date])
    elif specific_date:
        clauses.append("s.allocated_date = %s")
        params.append(specific_date)
    else:
        if month is None or year is None:
            try:
                month = int(request.args.get("month", today.month))
                year = int(request.args.get("year", today.year))
            except (TypeError, ValueError):
                month, year = today.month, today.year
        else:
            try:
                month = int(month)
                year = int(year)
            except (TypeError, ValueError):
                month, year = today.month, today.year
        clauses.append("s.allocated_date IS NOT NULL")
        clauses.append("MONTH(s.allocated_date) = %s")
        clauses.append("YEAR(s.allocated_date) = %s")
        params.extend([month, year])

    rows = run_query(
        f"""
        SELECT sh.client_id, sh.show_name, s.shot_code, s.allocated_date,
               s.mandays, s.client_feedback, s.status,
               s.supervisor_status, s.artist_status
        FROM shots s
        JOIN shows sh ON s.show_id = sh.show_id
        WHERE {' AND '.join(clauses)}
        ORDER BY s.allocated_date DESC, sh.client_id
        """,
        tuple(params),
        fetch_all=True,
    ) or []
    return rows


@reports_bp.route("", methods=["GET"])
@reports_bp.route("/", methods=["GET"])
@token_required
def report(_current_user_id):
    department = request.args.get("department")
    if not department:
        return failure("department is required.", 400)

    specific_date = request.args.get("date")  # YYYY-MM-DD overrides month/year
    start_date, end_date = _parse_date_range(
        request.args.get("startDate"), request.args.get("endDate")
    )

    rows = _report_rows(
        department,
        specific_date=specific_date,
        start_date=start_date,
        end_date=end_date,
    )

    items = [
        {
            "clientNo": r["client_id"],
            "show": r["show_name"],
            "shotId": r["shot_code"],
            "date": to_iso(r["allocated_date"]),
            "mandays": float(r["mandays"]) if r["mandays"] is not None else 0.0,
            "clientFeedback": r["client_feedback"],
            "progress": _progress_of(r),
        }
        for r in rows
    ]
    return success({"department": department, "items": items})


@reports_bp.route("/export", methods=["POST"])
@token_required
def export_report(_current_user_id):
    data = request.get_json(silent=True) or {}
    department = data.get("department") or request.args.get("department")
    if not department:
        return failure("department is required.", 400)

    specific_date = data.get("date") or request.args.get("date")
    start_date, end_date = _parse_date_range(
        data.get("startDate") or request.args.get("startDate"),
        data.get("endDate") or request.args.get("endDate"),
    )

    rows = _report_rows(
        department,
        specific_date=specific_date,
        start_date=start_date,
        end_date=end_date,
    )

    os.makedirs(EXPORT_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ["Client No", "Show", "Shot", "Date", "Mandays", "Client Feedback"]
    for col, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header_text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["client_id"])
        ws.cell(row=row_idx, column=2, value=row["show_name"])
        ws.cell(row=row_idx, column=3, value=row["shot_code"])
        ws.cell(row=row_idx, column=4, value=to_iso(row["allocated_date"]))
        ws.cell(
            row=row_idx,
            column=5,
            value=float(row["mandays"]) if row["mandays"] is not None else 0.0,
        )
        ws.cell(row=row_idx, column=6, value=row["client_feedback"])

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row_cells:
            cell.alignment = _CELL_ALIGN

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{max(ws.max_row, 1)}"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40

    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f"report_{_safe(department)}_{stamp}.xlsx"
    file_path = os.path.join(EXPORT_DIR, file_name)
    wb.save(file_path)

    return success(
        {
            "message": "Report exported.",
            "fileName": file_name,
            "savedPath": file_path,
            "downloadPath": f"/api/reports/export/download?fileName={file_name}",
        },
        201,
    )


@reports_bp.route("/export/download", methods=["GET"])
@token_required
def download_report_export(_current_user_id):
    file_name = (request.args.get("fileName") or "").strip()
    if not file_name:
        return failure("fileName is required.", 400)

    safe_name = os.path.basename(file_name)
    file_path = os.path.join(EXPORT_DIR, safe_name)
    if not os.path.isfile(file_path):
        return failure("Export file not found.", 404)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=safe_name,
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
