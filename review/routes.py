"""Review routes.

  * Department Review  -> monthly achievement: total shows, shots, mandays.
  * Individual Review  -> per-artist: shots worked + mandays delivered in a month.
  * Export             -> save each review as an .xlsx file under exports/reviews/.
"""

import os
import re
from datetime import date, datetime

from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from auth.middleware import token_required
from common.db_utils import get_user, run_query
from common.http import failure, success

review_bp = Blueprint("review", __name__)

# Folder on the backend server where exported reviews are stored.
EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports", "reviews")

_MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

_HEADER_FILL = PatternFill("solid", fgColor="3EBA02")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_CELL_ALIGN = Alignment(vertical="center")


def _month_year():
    today = date.today()
    try:
        month = int(request.args.get("month", today.month))
    except (TypeError, ValueError):
        month = today.month
    try:
        year = int(request.args.get("year", today.year))
    except (TypeError, ValueError):
        year = today.year
    return month, year


def _parse_date_range(start_date_raw=None, end_date_raw=None):
    """Parse start/end date values and return (start_date, end_date) or (None, None)."""
    start_date = start_date_raw if start_date_raw is not None else request.args.get("startDate")
    end_date = end_date_raw if end_date_raw is not None else request.args.get("endDate")
    
    try:
        if start_date:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        if end_date:
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        if start_date and end_date and start_date > end_date:
            start_date, end_date = end_date, start_date
        return start_date, end_date
    except (ValueError, TypeError):
        return None, None


def _month_name(month):
    return _MONTHS[month - 1] if 1 <= month <= 12 else str(month)


def _safe(value):
    """Make a string safe to use in a filename."""
    return re.sub(r"[^A-Za-z0-9_-]+", "_", str(value)).strip("_") or "review"


def _write_xlsx(title, summary_rows, detail_headers, detail_rows, filename):
    """Write a themed review workbook with summary + shot details sheets."""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{title[:24]} Summary"

    header = ws.cell(row=1, column=1, value="Field")
    header.fill = _HEADER_FILL
    header.font = _HEADER_FONT
    header.alignment = _HEADER_ALIGN
    header2 = ws.cell(row=1, column=2, value="Value")
    header2.fill = _HEADER_FILL
    header2.font = _HEADER_FONT
    header2.alignment = _HEADER_ALIGN

    for idx, (label, value) in enumerate(summary_rows, start=2):
        ws.cell(row=idx, column=1, value=label).alignment = _CELL_ALIGN
        ws.cell(row=idx, column=2, value=value).alignment = _CELL_ALIGN

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36
    ws.freeze_panes = "A2"

    details = wb.create_sheet(title=f"{title[:23]} Details")
    for col_idx, header_text in enumerate(detail_headers, start=1):
        cell = details.cell(row=1, column=col_idx, value=header_text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    for row_idx, row in enumerate(detail_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            details.cell(row=row_idx, column=col_idx, value=value).alignment = _CELL_ALIGN

    details.freeze_panes = "A2"
    if detail_headers:
        end_col = chr(ord("A") + len(detail_headers) - 1)
        details.auto_filter.ref = f"A1:{end_col}{max(details.max_row, 1)}"

    for col_idx, header_text in enumerate(detail_headers, start=1):
        width = max(14, min(42, len(header_text) + 6))
        details.column_dimensions[chr(ord("A") + col_idx - 1)].width = width

    if len(detail_headers) >= 7:
        details.column_dimensions["B"].width = 22
        details.column_dimensions["C"].width = 28
        details.column_dimensions["D"].width = 20
        details.column_dimensions["E"].width = 14
        details.column_dimensions["F"].width = 14
        details.column_dimensions["G"].width = 40

    path = os.path.join(EXPORT_DIR, filename)
    wb.save(path)
    return path



@review_bp.route("/department", methods=["GET"])
@token_required
def department_review(_current_user_id):
    department = request.args.get("department")
    if not department:
        return failure("department is required.", 400)
    
    # Check for date range params
    start_date, end_date = _parse_date_range()
    
    if start_date and end_date:
        # Date range filtering
        row = run_query(
            """
            SELECT COUNT(DISTINCT s.show_id) AS total_shows,
                   COUNT(*)                  AS total_shots,
                   COALESCE(SUM(s.mandays), 0) AS total_mandays
            FROM shots s
            WHERE s.department = %s
              AND s.allocated_date IS NOT NULL
              AND DATE(s.allocated_date) BETWEEN %s AND %s
            """,
            (department, start_date, end_date),
            fetch_one=True,
        )
        shot_rows = run_query(
            """
            SELECT sh.client_id,
                   sh.show_name,
                   s.shot_code,
                   s.allocated_date,
                   COALESCE(s.mandays, 0) AS mandays,
                   COALESCE(u.name, '') AS artist_name,
                   COALESCE(s.client_feedback, '') AS client_feedback
            FROM shots s
            JOIN shows sh ON sh.show_id = s.show_id
            LEFT JOIN users u ON u.user_id = s.artist_id
            WHERE s.department = %s
              AND s.allocated_date IS NOT NULL
              AND DATE(s.allocated_date) BETWEEN %s AND %s
            ORDER BY s.allocated_date DESC, sh.client_id, s.shot_code
            """,
            (department, start_date, end_date),
            fetch_all=True,
        ) or []
        return success(
            {
                "department": department,
                "startDate": str(start_date),
                "endDate": str(end_date),
                "totalShows": int(row["total_shows"]) if row else 0,
                "totalShots": int(row["total_shots"]) if row else 0,
                "totalMandays": float(row["total_mandays"]) if row else 0.0,
                "detailRows": [
                    {
                        "clientNo": r["client_id"],
                        "show": r["show_name"],
                        "shot": r["shot_code"],
                        "date": r["allocated_date"].strftime("%Y-%m-%d") if r.get("allocated_date") else None,
                        "mandays": float(r["mandays"]),
                        "artist": r["artist_name"],
                        "clientFeedback": r["client_feedback"],
                    }
                    for r in shot_rows
                ],
            }
        )
    else:
        # Month/year filtering (backward compatibility)
        month, year = _month_year()
        row = run_query(
            """
            SELECT COUNT(DISTINCT s.show_id) AS total_shows,
                   COUNT(*)                  AS total_shots,
                   COALESCE(SUM(s.mandays), 0) AS total_mandays
            FROM shots s
            WHERE s.department = %s
              AND s.allocated_date IS NOT NULL
              AND MONTH(s.allocated_date) = %s
              AND YEAR(s.allocated_date) = %s
            """,
            (department, month, year),
            fetch_one=True,
        )
        shot_rows = run_query(
            """
            SELECT sh.client_id,
                   sh.show_name,
                   s.shot_code,
                   s.allocated_date,
                   COALESCE(s.mandays, 0) AS mandays,
                   COALESCE(u.name, '') AS artist_name,
                   COALESCE(s.client_feedback, '') AS client_feedback
            FROM shots s
            JOIN shows sh ON sh.show_id = s.show_id
            LEFT JOIN users u ON u.user_id = s.artist_id
            WHERE s.department = %s
              AND s.allocated_date IS NOT NULL
              AND MONTH(s.allocated_date) = %s
              AND YEAR(s.allocated_date) = %s
            ORDER BY s.allocated_date DESC, sh.client_id, s.shot_code
            """,
            (department, month, year),
            fetch_all=True,
        ) or []
        return success(
            {
                "department": department,
                "month": month,
                "year": year,
                "totalShows": int(row["total_shows"]) if row else 0,
                "totalShots": int(row["total_shots"]) if row else 0,
                "totalMandays": float(row["total_mandays"]) if row else 0.0,
                "detailRows": [
                    {
                        "clientNo": r["client_id"],
                        "show": r["show_name"],
                        "shot": r["shot_code"],
                        "date": r["allocated_date"].strftime("%Y-%m-%d") if r.get("allocated_date") else None,
                        "mandays": float(r["mandays"]),
                        "artist": r["artist_name"],
                        "clientFeedback": r["client_feedback"],
                    }
                    for r in shot_rows
                ],
            }
        )


@review_bp.route("/individual", methods=["GET"])
@token_required
def individual_review(current_user_id):
    user_id = request.args.get("userId") or current_user_id
    user = get_user(user_id)
    if not user:
        return failure("User not found", 404)
    
    # Check for date range params
    start_date, end_date = _parse_date_range()
    
    if start_date and end_date:
        # Date range filtering
        row = run_query(
            """
            SELECT COUNT(*)                  AS shots_worked,
                   COALESCE(SUM(mandays), 0) AS mandays_delivered
            FROM shots
            WHERE artist_id = %s
              AND allocated_date IS NOT NULL
              AND DATE(allocated_date) BETWEEN %s AND %s
            """,
            (user_id, start_date, end_date),
            fetch_one=True,
        )
        shot_rows = run_query(
            """
            SELECT sh.client_id,
                   sh.show_name,
                   s.shot_code,
                   s.allocated_date,
                   COALESCE(s.mandays, 0) AS mandays,
                   COALESCE(s.artist_status, '') AS artist_status,
                   COALESCE(s.client_feedback, '') AS client_feedback
            FROM shots s
            JOIN shows sh ON sh.show_id = s.show_id
            WHERE s.artist_id = %s
              AND s.allocated_date IS NOT NULL
              AND DATE(s.allocated_date) BETWEEN %s AND %s
            ORDER BY s.allocated_date DESC, sh.client_id, s.shot_code
            """,
            (user_id, start_date, end_date),
            fetch_all=True,
        ) or []
        return success(
            {
                "userId": user_id,
                "name": user["name"],
                "department": user["department"],
                "startDate": str(start_date),
                "endDate": str(end_date),
                "shotsWorked": int(row["shots_worked"]) if row else 0,
                "mandaysDelivered": float(row["mandays_delivered"]) if row else 0.0,
                "detailRows": [
                    {
                        "clientNo": r["client_id"],
                        "show": r["show_name"],
                        "shot": r["shot_code"],
                        "date": r["allocated_date"].strftime("%Y-%m-%d") if r.get("allocated_date") else None,
                        "mandays": float(r["mandays"]),
                        "artistStatus": r["artist_status"],
                        "clientFeedback": r["client_feedback"],
                    }
                    for r in shot_rows
                ],
            }
        )
    else:
        # Month/year filtering (backward compatibility)
        month, year = _month_year()
        row = run_query(
            """
            SELECT COUNT(*)                  AS shots_worked,
                   COALESCE(SUM(mandays), 0) AS mandays_delivered
            FROM shots
            WHERE artist_id = %s
              AND allocated_date IS NOT NULL
              AND MONTH(allocated_date) = %s
              AND YEAR(allocated_date) = %s
            """,
            (user_id, month, year),
            fetch_one=True,
        )
        shot_rows = run_query(
            """
            SELECT sh.client_id,
                   sh.show_name,
                   s.shot_code,
                   s.allocated_date,
                   COALESCE(s.mandays, 0) AS mandays,
                   COALESCE(s.artist_status, '') AS artist_status,
                   COALESCE(s.client_feedback, '') AS client_feedback
            FROM shots s
            JOIN shows sh ON sh.show_id = s.show_id
            WHERE s.artist_id = %s
              AND s.allocated_date IS NOT NULL
              AND MONTH(s.allocated_date) = %s
              AND YEAR(s.allocated_date) = %s
            ORDER BY s.allocated_date DESC, sh.client_id, s.shot_code
            """,
            (user_id, month, year),
            fetch_all=True,
        ) or []
        return success(
            {
                "userId": user_id,
                "name": user["name"],
                "department": user["department"],
                "month": month,
                "year": year,
                "shotsWorked": int(row["shots_worked"]) if row else 0,
                "mandaysDelivered": float(row["mandays_delivered"]) if row else 0.0,
                "detailRows": [
                    {
                        "clientNo": r["client_id"],
                        "show": r["show_name"],
                        "shot": r["shot_code"],
                        "date": r["allocated_date"].strftime("%Y-%m-%d") if r.get("allocated_date") else None,
                        "mandays": float(r["mandays"]),
                        "artistStatus": r["artist_status"],
                        "clientFeedback": r["client_feedback"],
                    }
                    for r in shot_rows
                ],
            }
        )


@review_bp.route("/department/export", methods=["POST"])
@token_required
def export_department_review(_current_user_id):
    """Generate an .xlsx for the department review and save it on the server."""
    data = request.get_json(silent=True) or {}
    department = data.get("department") or request.args.get("department")
    if not department:
        return failure("department is required.", 400)
    start_date, end_date = _parse_date_range(
        data.get("startDate") or request.args.get("startDate"),
        data.get("endDate") or request.args.get("endDate"),
    )
    month, year = _month_year()

    if start_date and end_date:
        row = run_query(
            """
            SELECT COUNT(DISTINCT s.show_id) AS total_shows,
                         COUNT(*)                  AS total_shots,
                         COALESCE(SUM(s.mandays), 0) AS total_mandays
            FROM shots s
            WHERE s.department = %s
                AND s.allocated_date IS NOT NULL
                AND DATE(s.allocated_date) BETWEEN %s AND %s
            """,
            (department, start_date, end_date),
            fetch_one=True,
        )

        shot_rows = run_query(
            """
            SELECT sh.client_id,
                         sh.show_name,
                         s.shot_code,
                         s.allocated_date,
                         COALESCE(s.department, '') AS department,
                         COALESCE(s.mandays, 0) AS mandays,
                         COALESCE(u.name, '')   AS artist_name,
                         COALESCE(s.artist_status, '') AS artist_status,
                         COALESCE(s.client_feedback, '') AS client_feedback
            FROM shots s
            JOIN shows sh ON sh.show_id = s.show_id
            LEFT JOIN users u ON u.user_id = s.artist_id
            WHERE s.department = %s
                AND s.allocated_date IS NOT NULL
                AND DATE(s.allocated_date) BETWEEN %s AND %s
            ORDER BY s.allocated_date DESC, sh.client_id, s.shot_code
            """,
            (department, start_date, end_date),
            fetch_all=True,
        ) or []
        period_label = f"{start_date} to {end_date}"
    else:
        row = run_query(
            """
            SELECT COUNT(DISTINCT s.show_id) AS total_shows,
                         COUNT(*)                  AS total_shots,
                         COALESCE(SUM(s.mandays), 0) AS total_mandays
            FROM shots s
            WHERE s.department = %s
                AND s.allocated_date IS NOT NULL
                AND MONTH(s.allocated_date) = %s
                AND YEAR(s.allocated_date) = %s
            """,
            (department, month, year),
            fetch_one=True,
        )

        shot_rows = run_query(
            """
            SELECT sh.client_id,
                         sh.show_name,
                         s.shot_code,
                         s.allocated_date,
                         COALESCE(s.department, '') AS department,
                         COALESCE(s.mandays, 0) AS mandays,
                         COALESCE(u.name, '')   AS artist_name,
                         COALESCE(s.artist_status, '') AS artist_status,
                         COALESCE(s.client_feedback, '') AS client_feedback
            FROM shots s
            JOIN shows sh ON sh.show_id = s.show_id
            LEFT JOIN users u ON u.user_id = s.artist_id
            WHERE s.department = %s
                AND s.allocated_date IS NOT NULL
                AND MONTH(s.allocated_date) = %s
                AND YEAR(s.allocated_date) = %s
            ORDER BY s.allocated_date DESC, sh.client_id, s.shot_code
            """,
            (department, month, year),
            fetch_all=True,
        ) or []
        period_label = f"{_month_name(month)} {year}"

    rows = [
        ("Report Type", "Department Review"),
        ("Department", department),
        ("Period", period_label),
        ("Total Shows", int(row["total_shows"]) if row else 0),
        ("Total Shots", int(row["total_shots"]) if row else 0),
        ("Total Mandays", float(row["total_mandays"]) if row else 0.0),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    detail_headers = [
        "Client No",
        "Show",
        "Shot",
        "Date",
        "Department",
        "Mandays",
        "Artist",
        "Artist Status",
        "Client Feedback",
    ]
    detail_rows = [
        [
            r["client_id"],
            r["show_name"],
            r["shot_code"],
            r["allocated_date"].strftime("%Y-%m-%d") if r.get("allocated_date") else "",
            r["department"],
            float(r["mandays"]),
            r["artist_name"],
            r["artist_status"],
            r["client_feedback"],
        ]
        for r in shot_rows
    ]
    filename = f"department_{_safe(department)}_{year}_{month:02d}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    path = _write_xlsx("Department Review", rows, detail_headers, detail_rows, filename)

    return success(
        {
            "message": "Department review exported.",
            "fileName": filename,
            "savedPath": path,
            "downloadPath": f"/api/review/department/export/download?fileName={filename}",
        },
        201,
    )


@review_bp.route("/individual/export", methods=["POST"])
@token_required
def export_individual_review(current_user_id):
    """Generate an .xlsx for the individual review and save it on the server."""
    data = request.get_json(silent=True) or {}
    user_id = data.get("userId") or request.args.get("userId") or current_user_id
    user = get_user(user_id)
    if not user:
        return failure("User not found", 404)
    start_date, end_date = _parse_date_range(
        data.get("startDate") or request.args.get("startDate"),
        data.get("endDate") or request.args.get("endDate"),
    )
    month, year = _month_year()

    if start_date and end_date:
        row = run_query(
            """
            SELECT COUNT(*)                  AS shots_worked,
                         COALESCE(SUM(mandays), 0) AS mandays_delivered
            FROM shots
            WHERE artist_id = %s
                AND allocated_date IS NOT NULL
                AND DATE(allocated_date) BETWEEN %s AND %s
            """,
            (user_id, start_date, end_date),
            fetch_one=True,
        )

        shot_rows = run_query(
            """
            SELECT sh.client_id,
                         sh.show_name,
                         s.shot_code,
                         s.allocated_date,
                         COALESCE(s.department, '') AS department,
                         COALESCE(s.mandays, 0) AS mandays,
                         COALESCE(u.name, '') AS artist_name,
                         COALESCE(s.artist_status, '') AS artist_status,
                         COALESCE(s.client_feedback, '') AS client_feedback
            FROM shots s
            JOIN shows sh ON sh.show_id = s.show_id
            LEFT JOIN users u ON u.user_id = s.artist_id
            WHERE s.artist_id = %s
                AND s.allocated_date IS NOT NULL
                AND DATE(s.allocated_date) BETWEEN %s AND %s
            ORDER BY s.allocated_date DESC, sh.client_id, s.shot_code
            """,
            (user_id, start_date, end_date),
            fetch_all=True,
        ) or []
        period_label = f"{start_date} to {end_date}"
    else:
        row = run_query(
            """
            SELECT COUNT(*)                  AS shots_worked,
                         COALESCE(SUM(mandays), 0) AS mandays_delivered
            FROM shots
            WHERE artist_id = %s
                AND allocated_date IS NOT NULL
                AND MONTH(allocated_date) = %s
                AND YEAR(allocated_date) = %s
            """,
            (user_id, month, year),
            fetch_one=True,
        )

        shot_rows = run_query(
            """
            SELECT sh.client_id,
                         sh.show_name,
                         s.shot_code,
                         s.allocated_date,
                         COALESCE(s.department, '') AS department,
                         COALESCE(s.mandays, 0) AS mandays,
                         COALESCE(u.name, '') AS artist_name,
                         COALESCE(s.artist_status, '') AS artist_status,
                         COALESCE(s.client_feedback, '') AS client_feedback
            FROM shots s
            JOIN shows sh ON sh.show_id = s.show_id
            LEFT JOIN users u ON u.user_id = s.artist_id
            WHERE s.artist_id = %s
                AND s.allocated_date IS NOT NULL
                AND MONTH(s.allocated_date) = %s
                AND YEAR(s.allocated_date) = %s
            ORDER BY s.allocated_date DESC, sh.client_id, s.shot_code
            """,
            (user_id, month, year),
            fetch_all=True,
        ) or []
        period_label = f"{_month_name(month)} {year}"

    rows = [
        ("Report Type", "Individual Review"),
        ("Name", user["name"]),
        ("Department", user["department"]),
        ("Period", period_label),
        ("Shots Worked", int(row["shots_worked"]) if row else 0),
        ("Mandays Delivered", float(row["mandays_delivered"]) if row else 0.0),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    detail_headers = [
        "Client No",
        "Show",
        "Shot",
        "Date",
        "Department",
        "Mandays",
        "Artist",
        "Artist Status",
        "Client Feedback",
    ]
    detail_rows = [
        [
            r["client_id"],
            r["show_name"],
            r["shot_code"],
            r["allocated_date"].strftime("%Y-%m-%d") if r.get("allocated_date") else "",
            r["department"],
            float(r["mandays"]),
            r["artist_name"],
            r["artist_status"],
            r["client_feedback"],
        ]
        for r in shot_rows
    ]
    filename = f"individual_{_safe(user['name'])}_{year}_{month:02d}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    path = _write_xlsx("Individual Review", rows, detail_headers, detail_rows, filename)

    return success(
        {
            "message": "Individual review exported.",
            "fileName": filename,
            "savedPath": path,
            "downloadPath": f"/api/review/individual/export/download?fileName={filename}",
        },
        201,
    )


@review_bp.route("/department/export/download", methods=["GET"])
@token_required
def download_department_export(_current_user_id):
    file_name = (request.args.get("fileName") or "").strip()
    if not file_name:
        return failure("fileName is required.", 400)

    safe_name = os.path.basename(file_name)
    file_path = os.path.join(EXPORT_DIR, safe_name)
    if not os.path.isfile(file_path):
        return failure("Export file not found.", 404)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=safe_name,
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


@review_bp.route("/individual/export/download", methods=["GET"])
@token_required
def download_individual_export(_current_user_id):
    file_name = (request.args.get("fileName") or "").strip()
    if not file_name:
        return failure("fileName is required.", 400)

    safe_name = os.path.basename(file_name)
    file_path = os.path.join(EXPORT_DIR, safe_name)
    if not os.path.isfile(file_path):
        return failure("Export file not found.", 404)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=safe_name,
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
